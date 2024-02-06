import openpyxl
import csv
import random
import string
import os
import shutil
import zipfile
from django.conf import settings
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile


def create_backup(full_file_path):
    backup_path = full_file_path.replace('.xlsx', '_backup.xlsx')
    shutil.copyfile(full_file_path, backup_path)
    return backup_path


def zip_generated_files():
    directory_path = os.path.join(settings.MEDIA_ROOT, 'generated_files')
    zip_file_path = os.path.join(settings.MEDIA_ROOT, 'generated_files.zip')

    with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(directory_path):
            for file in files:
                file_path = os.path.join(root, file)
                zipf.write(file_path, os.path.relpath(file_path, os.path.join(directory_path, '..')))

    return zip_file_path


def generate_random_password(length=8, special_chars="!#?", special_count=2):
    special_indices = random.sample(range(length), special_count)
    characters = string.ascii_letters + string.digits
    password_list = [random.choice(characters) for _ in range(length)]

    for index in special_indices:
        password_list[index] = random.choice(special_chars)

    return ''.join(password_list)


def generate_passwords_and_save_to_csv(full_file_path, file_name):
    output_messages = []
    base_name = os.path.splitext(os.path.basename(full_file_path))[0]
    directory = os.path.join(settings.MEDIA_ROOT, 'generated_files')
    os.makedirs(directory, exist_ok=True)

    workbook = openpyxl.load_workbook(full_file_path)
    csv_file_paths = []

    for sheet_name in workbook.sheetnames:
        csv_file_path = os.path.join(directory, f"{base_name}_{sheet_name}_passwords.csv")
        csv_file_paths.append(csv_file_path)

        with open(csv_file_path, 'w', newline='') as csv_file:
            writer = csv.writer(csv_file)
            sheet = workbook[sheet_name]
            nutzer_column = None
            passwort_column = None

            for i, column in enumerate(sheet.iter_cols(min_row=1, max_row=1)):
                if column[0].value == "Nutzer":
                    nutzer_column = i
                elif column[0].value == "Passwort":
                    passwort_column = i

            if nutzer_column is None or passwort_column is None:
                output_messages.append(f"Spalte 'Nutzer' oder 'Passwort' in '{sheet_name}' nicht gefunden.\n")
                continue

            base_password = generate_random_password()
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                nutzer_value = row[nutzer_column].value
                if nutzer_value:
                    nutzer_nummer = ''.join(filter(str.isdigit, nutzer_value))
                    password = base_password + nutzer_nummer
                    row[passwort_column].value = password

                    nutzer_simple = nutzer_value.split('@')[0]
                    writer.writerow([nutzer_simple, password])
                    output_messages.append(f"Passwort für {nutzer_simple} in '{sheet_name}' gesetzt: {password}\n")

    workbook.save(full_file_path)  # Speichern der bearbeiteten Excel-Datei
    output_messages.append("Passwörter erfolgreich generiert und Excel-Datei gespeichert.\n")

    return full_file_path, csv_file_paths, '\n'.join(output_messages)


def cleanup_generated_files():
    directory_path = os.path.join(settings.MEDIA_ROOT, 'generated_files')
    # Sicherstellen, dass der Ordner existiert
    if os.path.exists(directory_path):
        # Löschen aller Dateien und Unterverzeichnisse im Ordner
        shutil.rmtree(directory_path)
        # Erneutes Erstellen des leeren Ordners
        os.makedirs(directory_path, exist_ok=True)
